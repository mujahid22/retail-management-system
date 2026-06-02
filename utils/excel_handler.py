"""Read and write the star schema XLSX file."""
import os
from pathlib import Path
from typing import Dict

import pandas as pd
from openpyxl import load_workbook

DATA_PATH = Path(__file__).parent.parent / "data" / "sales_data.xlsx"
SHEET_ORDER = ["fact_sales", "dim_customer", "dim_product", "dim_date", "dim_geography"]


def ensure_data_exists() -> None:
    """Generate sample data if the XLSX does not exist yet."""
    if not DATA_PATH.exists():
        DATA_PATH.parent.mkdir(parents=True, exist_ok=True)
        from utils.data_generator import generate_all
        generate_all(str(DATA_PATH))


def load_all() -> Dict[str, pd.DataFrame]:
    ensure_data_exists()
    return pd.read_excel(DATA_PATH, sheet_name=None, engine="openpyxl")


def load_sheet(sheet: str) -> pd.DataFrame:
    ensure_data_exists()
    return pd.read_excel(DATA_PATH, sheet_name=sheet, engine="openpyxl")


def _next_id(df: pd.DataFrame, id_col: str, prefix: str, width: int = 8) -> str:
    """Return the next sequential ID based on current row count."""
    return f"{prefix}{str(len(df) + 1).zfill(width)}"


def upsert_customer(data: dict) -> str:
    """Insert or return existing customer_id."""
    df = load_sheet("dim_customer")
    match = df[df["email"].str.lower() == data["email"].lower()]
    if not match.empty:
        return match.iloc[0]["customer_id"]
    new_id = _next_id(df, "customer_id", "C", 6)
    new_row = {
        "customer_id": new_id,
        "customer_name": data["customer_name"],
        "email": data["email"],
        "phone": data["phone"],
        "age_group": data["age_group"],
        "gender": data["gender"],
    }
    _append_row("dim_customer", new_row)
    return new_id


def upsert_geography(data: dict) -> str:
    """Insert or return existing geography_id."""
    df = load_sheet("dim_geography")
    mask = (df["city"].str.lower() == data["city"].lower()) & \
           (df["state"].str.upper() == data["state"].upper())
    match = df[mask]
    if not match.empty:
        return match.iloc[0]["geography_id"]
    new_id = _next_id(df, "geography_id", "G", 4)
    from utils.schema import US_REGIONS
    region = "Other"
    for r, states in US_REGIONS.items():
        if data["state"].upper() in states:
            region = r
            break
    new_row = {
        "geography_id": new_id,
        "city": data["city"],
        "state": data["state"].upper(),
        "region": region,
        "zip_code": data.get("zip_code", "00000"),
    }
    _append_row("dim_geography", new_row)
    return new_id


def ensure_date_exists(d) -> None:
    """Insert d into dim_date if it is not already present."""
    from datetime import date as date_type
    import datetime as _dt
    if not isinstance(d, date_type):
        d = pd.Timestamp(d).date()
    date_id = int(d.strftime("%Y%m%d"))
    df = load_sheet("dim_date")
    if date_id in df["date_id"].values:
        return
    iso = d.isocalendar()
    new_row = {
        "date_id": date_id,
        "full_date": d,
        "year": d.year,
        "quarter": (d.month - 1) // 3 + 1,
        "month": d.month,
        "month_name": d.strftime("%B"),
        "week_of_year": iso[1],
        "day_of_week": d.weekday(),
        "day_name": d.strftime("%A"),
        "is_weekend": d.weekday() >= 5,
    }
    _append_row("dim_date", new_row)


def append_sale_rows(rows: list[dict]) -> None:
    """Append multiple fact_sales rows at once."""
    ensure_data_exists()
    wb = load_workbook(DATA_PATH)
    ws = wb["fact_sales"]
    for row in rows:
        ws.append([
            row["sale_id"], row["transaction_id"], row["customer_id"],
            row["product_id"], row["date_id"], row["geography_id"],
            row["qty"], row["unit_cost"], row["unit_price"],
            row["discount_pct"], row["discount_amount"],
            row["subtotal"], row["tax_rate"], row["tax_amount"],
            row["total_amount"], row["payment_mode"],
        ])
    wb.save(DATA_PATH)


def _append_row(sheet: str, row: dict) -> None:
    ensure_data_exists()
    wb = load_workbook(DATA_PATH)
    ws = wb[sheet]
    ws.append(list(row.values()))
    wb.save(DATA_PATH)


def next_sale_id() -> str:
    df = load_sheet("fact_sales")
    return _next_id(df, "sale_id", "S", 8)
